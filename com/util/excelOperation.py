# -*- coding: utf-8 -*-
"""
@Time ： 2022/5/21 13:03
@Auth ： muzili
@File ： excelOperation.py 暂时用不到该文件
@IDE  ： PyCharm
"""
import openpyxl
import xlrd
import xlwt


class ReadXLS(object):
    """读取xls文件"""

    def __init__(self, filename, sheet_name='Sheet1'):
        """
        :param filename: 文件名
        :param sheet_name: sheet名
        """
        self.work_book = xlrd.open_workbook(filename)
        self.sheet_name = self.work_book.sheet_by_name(sheet_name)

    def get_rows(self, row_x=0, start_col=0, end_col=None):
        """
        获取某一行的所有值
        :param row_x: 第几行
        :param start_col: 默认从第0列开始读取
        :param end_col: 默认读取所有列
        :return:
        """
        if row_x > self.sheet_name.nrows:
            return None
        return self.sheet_name.row_values(row_x, start_colx=start_col, end_colx=end_col)

    def get_cols(self, col_x, start_row=0, end_row=None):
        """
        获取某一列的所有值
        :param col_x:
        :param start_row: 默认从第0行开始读取 从0行开始计算
        :param end_row: 默认读取所有行
        :return:
        """
        if col_x > self.sheet_name.ncols or start_row > self.sheet_name.nrows:
            return None
        return self.sheet_name.col_values(col_x, start_rowx=start_row, end_rowx=end_row)

    def get_cell(self, row_x=0, col_x=0):
        """
        获取某个单元格
        :param row_x:
        :param col_x:
        :return:
        """
        if row_x > self.sheet_name.nrows or col_x > self.sheet_name.ncols:
            return None
        return self.sheet_name.cell_value(rowx=row_x, colx=col_x)

    def get_row_index(self, col_x=0, value=None):
        """
        根据value获取value的行号
        :param col_x:
        :param value:
        :return:
        """
        col_values = self.get_cols(col_x)
        row_index = list()
        for i in range(len(col_values)):
            if value == col_values[i]:
                row_index.append(i)
        return row_index


class WriteXLS(object):
    """写.xls文件"""

    def __init__(self, filename, sheet_name='Sheet1'):
        pass

    def post_cell(self, row_x=1, col_x=15, value='test12'):
        pass


class ReadExcel(object):
    """读取.xlsx文件"""

    def __init__(self, filename, sheet_name='Sheet1'):
        """
        :param filename: 文件名
        :param sheet_name: sheet名
        """
        self.work_book = openpyxl.load_workbook(filename, read_only=True)
        self.sheet_name = self.work_book[sheet_name]

    def get_rows(self, row_x, start_col=1, end_col=None):
        """
        获取某一行的所有值
        :param row_x:
        :param start_col: 默认从第0列开始读取
        :param end_col: 默认读取所有列
        :return:
        """
        values = list()
        cols = self.sheet_name.max_column
        if end_col is None:
            end_col = cols
        if start_col > cols or start_col > end_col:
            return values
        row_x += 1  # 为了保持与 ReadXLS 的 get_rows 函数一致，这里需要加 1
        end_col += 1
        for col in range(start_col, end_col):
            values.append(self.sheet_name.cell(row=row_x, column=col).value)
        return values

    def get_cols(self, col_x, start_row=1, end_row=None):
        """ ['用例编号', '', '', '', '', '', '', '', '', 'tesst_gushiwen.py']
        获取某一列的所有值
        :param col_x:
        :param start_row: 默认从第1行开始读取 从1行开始计算
        :param end_row: 默认读取所有行
        :return:
        """
        values = list()
        rows = self.sheet_name.max_row
        if end_row is None:
            end_row = rows
        if start_row > rows or start_row > end_row:
            return values
        col_x += 1  # 为了保持与 ReadXLS 的 get_cols 函数一致，这里需要加 1
        end_row += 1
        for row in range(start_row, end_row):
            values.append(self.sheet_name.cell(row=row, column=col_x).value)
        return values

    def get_cell(self, row_x=1, col_x=1):
        """
        获取某个单元格
        :param row_x:
        :param col_x:
        :return:
        """
        row_x += 1  # 为了保持与 ReadXLS 的 get_cell 函数一致，这里需要加 1
        col_x += 1
        return self.sheet_name.cell(row=row_x, column=col_x).value

    def get_row_index(self, col_x=0, value=''):
        """
        根据value获取value的行号
        :param col_x:
        :param value:
        :return:
        """
        col_values = self.get_cols(col_x)
        row_index = list()
        for i in range(len(col_values)):
            if value == col_values[i]:
                row_index.append(i)
        return row_index


class WriteExcel(object):
    """写.xlsx文件"""

    def __init__(self, filename, sheet_name='Sheet1'):
        """
        :param filename: 文件名
        :param sheet_name: sheet名
        """
        self.filename = filename
        self.work_book = openpyxl.load_workbook(filename)
        self.sheet_name = self.work_book[sheet_name]

    def post_cell(self, row_x=2, col_x=16, value='test12'):
        """
        写cell单元的值，默认写实际结果的值
        :param row_x:
        :param col_x:
        :param value:
        :return:
        """
        cols = self.sheet_name.max_column
        rows = self.sheet_name.max_row
        if row_x > rows or col_x > cols:
            raise KeyError("写入行位置大于最大行数or写入列位置大于最大列数")
        row_x += 1  # 为了保持与 WriteXLS 的 post_cell 函数一致，这里需要加 1
        col_x += 1
        val = self.sheet_name.cell(row=row_x, column=col_x).value
        if val is None:
            val = 'auto write:\n'
        elif 'auto write' not in val:
            val = val + '\nauto write:\n'
        else:
            val = val + '\n'
        self.sheet_name.cell(row=row_x, column=col_x).value = val + value
        self.work_book.save(self.filename)


if __name__ == "__main__":
    pass
